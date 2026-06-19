from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import DEFAULT_THEME, EXPORT_SHEET_NAME


THEMES = {
    "Classic Navy": {
        "header_fill": "1F497D",
        "header_text": "FFFFFF",
        "zebra_fill": "F2F5F9",
    },
    "Emerald Green": {
        "header_fill": "0E6251",
        "header_text": "FFFFFF",
        "zebra_fill": "E8F8F5",
    },
    "Charcoal Tech": {
        "header_fill": "2C3E50",
        "header_text": "FFFFFF",
        "zebra_fill": "F4F6F7",
    },
}


def apply_excel_theme(writer, df, theme_name):
    """Apply a predefined workbook theme without relying on model output."""
    worksheet = writer.sheets[EXPORT_SHEET_NAME]
    selected_theme = THEMES.get(theme_name, THEMES[DEFAULT_THEME])

    header_font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color=selected_theme["header_text"],
    )
    header_fill = PatternFill(
        start_color=selected_theme["header_fill"],
        end_color=selected_theme["header_fill"],
        fill_type="solid",
    )
    data_font = Font(name="Calibri", size=11, bold=False)
    zebra_fill = PatternFill(
        start_color=selected_theme["zebra_fill"],
        end_color=selected_theme["zebra_fill"],
        fill_type="solid",
    )
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for column_index in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_index in range(2, len(df) + 2):
        is_even_row = row_index % 2 == 0
        for column_index in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.font = data_font
            cell.border = thin_border
            if is_even_row:
                cell.fill = zebra_fill

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        column_letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[column_letter].width = max(max_length + 3, 12)
